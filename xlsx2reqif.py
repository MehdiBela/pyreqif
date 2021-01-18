#!/usr/bin/env python3

"""
Main module of the project
Usage: ./xlsx2reqif.py source.xlsx
"""

import os
import io
import sys
import openpyxl
import pyreqif.create
import pyreqif.reqif

if __name__ == "__main__":
    file_name = sys.argv[1]
    document_title, _ = os.path.splitext(os.path.basename(file_name))
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    cols = []
    for col_nr in range(1, ws.max_column + 1):
        cols.append(ws.cell(1, col_nr).value)

    DOCUMENT_REQIF_ID = "_{}ReqifId-Header".format(document_title)
    SPEC_REQIF_ID = "_{}ReqifId--spec".format(document_title)
    DOC_TYPE_REF = "_DOC_TYPE_REF"

    # create doc:
    document = pyreqif.create.createDocument(DOCUMENT_REQIF_ID, title=document_title)
    # pyreqif.create.addDocType(DOC_TYPE_REF, document)

    # create primitive datatype
    pyreqif.create.addDatatype("_datatype_ID", document, longName=None)

    # create columns
    SPEC_ID = "_some_requirement_type_id"
    SPEC_LONG_NAME = "Requirement attributes"
    TYPE_REF = "_datatype_ID"
    for col in cols:
        pyreqif.create.addReqType(
            SPEC_ID, SPEC_LONG_NAME, "_reqtype_for_" + col.replace(" ", "_"),
            col, "_datatype_ID", document
        )

    # create document hierarchy head
    hierarchy = pyreqif.create.createHierarchHead(
        document_title, typeRef=DOC_TYPE_REF, id=SPEC_REQIF_ID
    )

    # create child elements
    hierarchy_stack = []
    last_hierarchy_element = hierarchy
    for row_nr in range(2, ws.max_row + 1):
        xls_req = dict(zip(cols, [ws.cell(row_nr, x).value for x in range(1, ws.max_column + 1)]))
        if "reqifId" not in xls_req:
            xls_req["reqifId"] = pyreqif.create.creatUUID()
        for col in cols:
            if isinstance(xls_req[col], str):
                xls_req[col] = xls_req[col].replace("<", "&gt;")
                xls_req[col].replace("<", "&lt;")
            if xls_req[col] is not None:
                CONTENT = "<div>%s</div>" % str(xls_req[col])
                REQ_TYPE_REF = "_reqtype_for_" + col.replace(" ", "_")
                pyreqif.create.addReq(
                    xls_req["reqifId"], SPEC_ID, CONTENT, REQ_TYPE_REF, document
                )

        # do hierarchy
        hierarchy_element = pyreqif.create.createHierarchElement(xls_req["reqifId"])
        level = ws.row_dimensions[row_nr].outline_level + 1
        if level > len(hierarchy_stack):
            hierarchy_stack.append(last_hierarchy_element)
        elif level < len(hierarchy_stack):
            hierarchy_stack = hierarchy_stack[0:level]

        current_head = hierarchy_stack[level - 1]
        current_head.addChild(hierarchy_element)
        last_hierarchy_element = hierarchy_element

    document.hierarchy.append(hierarchy)

    # save reqif to string
    strIO = io.StringIO()
    pyreqif.reqif.dump(document, strIO)
    with open(document_title + '.reqif', "w") as f:
        f.write(strIO.getvalue())
