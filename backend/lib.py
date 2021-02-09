import configparser

import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

REQIF_LONG_NAME = "ReqIF.LongName"


def get_excel_preview_data(file_bytes):
    """
    Return the two first line of the provided excel file
    Will throw an error in the file has more than one sheet
    """
    wb = openpyxl.load_workbook(file_bytes)
    sheets_count = len(wb._sheets)
    if sheets_count > 1:
        raise TooManySheetsException(sheets_count)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=1, max_row=2):
        data.append([cell.value for cell in row])
    return data


def reorder_data(file_bytes, headers, removed_columns):
    """
    Create a new workbook from original data according to user defined column orders
    """
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb.active
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.insert_cols(1, len(headers))
    for i in removed_columns:
        ws.delete_cols(i + 1)
    if REQIF_LONG_NAME in headers:
        # long name has to be first column
        idx = headers.index(REQIF_LONG_NAME)

        for i, row in enumerate(ws.rows):
            new_cell = new_ws.cell(i + 1, 1)
            new_cell.value = row[idx].value
    for header in headers:
        idx = headers.index(header)
        if header == REQIF_LONG_NAME:
            continue
        for i, row in enumerate(ws.rows):
            new_cell = new_ws.cell(i + 1, (idx + 2) if REQIF_LONG_NAME in headers else (idx + 1))
            value = row[idx].value
            if hasattr(row[idx], "number_format") and isinstance(value, int):
                new_cell.value = row[idx].number_format.replace("#", str(value))
            else:
                new_cell.value = value
    for i, col in enumerate(new_ws.iter_cols()):
        if not col[0].value:
            new_ws.delete_cols(i + 1)
    new_wb.save("out.xlsx")
    return save_virtual_workbook(new_wb)


def get_saved_configurations(configuration_file):
    config = configparser.ConfigParser()
    config.read(configuration_file)
    configurations = []
    for i in config.sections():
        headers = config[i].get("headers")
        headers = headers.split(",")
        removed_columns = config[i].get("removed_columns")
        removed_columns = [int(i) for i in removed_columns.split(",")] if removed_columns else []
        configurations.append({
            "name": i,
            "headers": headers,
            "removed_columns": removed_columns
        })
    return configurations


def add_configuration(name, headers, removed_columns, configuration_file):
    new_config = {
        "headers": ",".join(headers),
        "removed_columns": ",".join([str(i) for i in removed_columns])
    }
    config = configparser.ConfigParser()
    config.read(configuration_file)
    config[name] = new_config
    new_config["name"] = name
    with open(configuration_file, "w") as f:
        config.write(f)


class TooManySheetsException(Exception):
    """

    """

    def __init__(self, sheets_count, message="Too many sheets in file (%s)"):
        self.sheets_count = sheets_count
        self.message = message % self.sheets_count
        super().__init__(self.message)
