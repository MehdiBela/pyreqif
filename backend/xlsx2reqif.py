#!/usr/bin/env python3

"""
Main module of the project
Usage: ./xlsx2reqif.py source.xlsx
"""

import io
import openpyxl
import pyreqif.create
import pyreqif.reqif


def convert_file(excel_file, file_name, save_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    cols = []
    for col_nr in range(1, ws.max_column + 1):
        cols.append(ws.cell(1, col_nr).value)

    document_reqif_id = "_{}ReqifId-Header".format(file_name)
    spec_reqif_id = "_{}ReqifId--spec".format(file_name)
    doc_type_ref = "_DOC_TYPE_REF"

    # create doc:
    document = pyreqif.create.createDocument(document_reqif_id, title=file_name)
    # pyreqif.create.addDocType(doc_type_ref, document)

    # create primitive datatype
    pyreqif.create.addDatatype("_datatype_ID", document, longName=None)

    # create columns
    spec_id = "_some_requirement_type_id"
    spec_long_name = "Requirement attributes"
    type_ref = "_datatype_ID"
    for col in cols:
        pyreqif.create.addReqType(
            spec_id, spec_long_name, "_reqtype_for_" + col.replace(" ", "_"),
            col, "_datatype_ID", document
        )

    # create document hierarchy head
    hierarchy = pyreqif.create.createHierarchHead(
        file_name, typeRef=doc_type_ref, id=spec_reqif_id
    )

    # create child elements
    hierarchy_stack = []
    last_hierarchy_element = hierarchy
    for row_nr in range(2, ws.max_row + 1):
        xls_req = dict(zip(cols, [ws.cell(row_nr, x).value for x in range(1, ws.max_column + 1)]))
        if "reqifId" not in xls_req:
            xls_req["reqifId"] = pyreqif.create.creatUUID()
        for col in cols:
            if isinstance(xls_req[col], str):
                xls_req[col] = xls_req[col].replace("<", "&gt;")
                xls_req[col].replace("<", "&lt;")
            if xls_req[col] is not None:
                content = "<div>%s</div>" % str(xls_req[col])
                req_type_ref = "_reqtype_for_" + col.replace(" ", "_")
                pyreqif.create.addReq(
                    xls_req["reqifId"], spec_id, content, req_type_ref, document
                )

        # do hierarchy
        hierarchy_element = pyreqif.create.createHierarchElement(xls_req["reqifId"])
        level = ws.row_dimensions[row_nr].outline_level + 1
        if level > len(hierarchy_stack):
            hierarchy_stack.append(last_hierarchy_element)
        elif level < len(hierarchy_stack):
            hierarchy_stack = hierarchy_stack[0:level]

        current_head = hierarchy_stack[level - 1]
        current_head.addChild(hierarchy_element)
        last_hierarchy_element = hierarchy_element

    document.hierarchy.append(hierarchy)

    # save reqif to string
    str_io = io.StringIO()
    pyreqif.reqif.dump(document, str_io)
    if save_file:
        return str_io.getvalue()
    else:
        return str_io.getvalue()
